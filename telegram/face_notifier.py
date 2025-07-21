import pyodbc
import time
import httpx
import asyncio
import pandas as pd
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# SQL Server connection info
SQL_CONNECTION_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost;"
    "DATABASE=hik;"
    "UID=sa;"
    "PWD=Quvonchbek2006@"
)

API_BASE = "http://127.0.0.1:8000/api/v1"
TELEGRAM_BOT_TOKEN = "8061178459:AAFffd5n0Vd_NQIxz8dAdD2SbGEA-H9dxGQ"
EXCEL_SEND_IDS = [101927389, 5091336899]

# Track already sent messages per day
student_count = 0
sent_kirish = set()
sent_chiqish = set()
kirish_log = {}
teachers_student_count = {}
last_chiqish_sent_date = None
excel_sent_date = None
current_date = date.today()

# Excel save path
excel_folder = Path("kirish_logs")
excel_folder.mkdir(exist_ok=True)


# Get today's entries from database
def get_today_entries():
    conn = pyodbc.connect(SQL_CONNECTION_STRING)
    cursor = conn.cursor()
    today = datetime.now().date()
    cursor.execute("""
        SELECT employeeID, personName, authDateTime, deviceName
        FROM imvs
        WHERE CAST(authDateTime AS DATE) = ?
        ORDER BY authDateTime ASC
    """, today)
    rows = cursor.fetchall()
    conn.close()
    return rows


# Send Telegram message
async def send_telegram_message(student_id, name, timestamp, direction):
    async with httpx.AsyncClient() as client:
        resp = await client.get(f"{API_BASE}/get-telegram-id/", params={"student_id": str(student_id)})
        if resp.status_code != 200:
            print(f"❌ Telegram ID topilmadi: {student_id}")
            return

        telegram_id = resp.json()["telegram_id"]
        direction = direction.lower()

        if "kirish" in direction:
            status = "maktabga kirdi"
        elif "chiqish" in direction:
            status = "maktabdan chiqdi"
        else:
            status = f"{direction.lower()} bo‘ldi"
        arr = name.split()
        name = f"{arr[0]} {arr[1]}"
        msg = (
            f"📢 Farzandingiz <b>{name}</b> "
            f"{timestamp.strftime('%Y-%m-%d %H:%M:%S')} da {status}."
        )

        telegram_api = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        await client.post(telegram_api, json={
            "chat_id": telegram_id,
            "text": msg,
            "parse_mode": "HTML"
        })
        print(f"✅ Telegram ID topildi: {student_id}")


# Send Excel to specified Telegram users
async def send_excel_file(excel_path):
    async with httpx.AsyncClient() as client:
        for chat_id in EXCEL_SEND_IDS:
            with open(excel_path, "rb") as file:
                files = {'document': (
                    excel_path.name, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'chat_id': chat_id}
                telegram_api = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
                response = await client.post(telegram_api, data=data, files=files)
                print(f"📤 Excel sent to {chat_id} with status: {response.status_code}")


# Process entries
async def process_entries():
    global last_chiqish_sent_date, current_date, sent_kirish, sent_chiqish, kirish_log, excel_sent_date, student_count, teachers_student_count

    if date.today() != current_date:
        current_date = date.today()
        sent_kirish.clear()
        sent_chiqish.clear()
        kirish_log.clear()
        teachers_student_count.clear()
        last_chiqish_sent_date = None
        excel_sent_date = None
        student_count = 0

    rows = get_today_entries()
    chiqish_latest = {}

    for row in rows:
        emp_id = str(int(row.employeeID))
        direction = row.deviceName.lower()

        if "kirish" in direction and emp_id not in sent_kirish:
            sent_kirish.add(emp_id)
            student_name = row.personName.lower().split()
            if len(student_name) > 2:
                student_count += 1
            for i in range(2, len(student_name)):
                if teachers_student_count.get(student_name[i]):
                    teachers_student_count[student_name[i]] += 1
                else:
                    teachers_student_count[student_name[i]] = 1
            kirish_log[emp_id] = {
                "ID": emp_id,
                "Name": f"{student_name[0]} {student_name[1]}",
                "Entry Time": row.authDateTime.strftime('%Y-%m-%d %H:%M:%S')
            }
            await send_telegram_message(emp_id, row.personName, row.authDateTime, row.deviceName)

        if "chiqish" in direction:
            chiqish_latest[emp_id] = row

    now = datetime.now()
    today = now.date()

    # Send chiqish messages
    if (now.hour == 17 and 40 <= now.minute < 45 and last_chiqish_sent_date != today):
        for emp_id, row in chiqish_latest.items():
            if emp_id not in sent_chiqish:
                sent_chiqish.add(emp_id)
                await send_telegram_message(emp_id, row.personName, row.authDateTime, row.deviceName)
        last_chiqish_sent_date = today

    # Save and send Excel
    # Save and send Excel
    # Save and send Excel
    if now.hour == 10 and 5 <= now.minute < 10 and excel_sent_date != today and kirish_log:
        print("entered")
        # Prepare data frames
        file_name = f"{today.strftime('%Y-%m-%d')}.xlsx"
        file_path = excel_folder / file_name

        # Create a new Excel workbook
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        workbook = writer.book

        # Sheet 1: Summary Report
        summary_data = [
            ["Total Students", student_count],
            ["", ""],  # Empty row for spacing
            ["Teacher Summary", ""]
        ]

        # Add teacher summary
        for teacher, count in teachers_student_count.items():
            summary_data.append([teacher, count])

        summary_df = pd.DataFrame(summary_data, columns=["Category", "Count"])

        # Sheet 2: Detailed Entries
        entries_data = []
        for emp_id, data in kirish_log.items():
            entries_data.append([
                data["ID"],
                data["Name"],
                data["Entry Time"]
            ])

        entries_df = pd.DataFrame(entries_data, columns=["ID", "Name", "Entry Time"])

        # Write sheets to Excel
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        entries_df.to_excel(writer, sheet_name="Detailed Entries", index=False)

        # Get sheets for styling
        summary_sheet = writer.sheets["Summary"]
        entries_sheet = writer.sheets["Detailed Entries"]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        total_fill = PatternFill("solid", fgColor="FFC000")  # Orange for totals
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Style Summary sheet
        # Headers
        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Total row
        summary_sheet['A1'].fill = total_fill
        summary_sheet['B1'].fill = total_fill

        # Teacher summary header
        summary_sheet['A3'].fill = header_fill
        summary_sheet['A3'].font = header_font
        summary_sheet['B3'].fill = header_fill
        summary_sheet['B3'].font = header_font

        # Style all cells
        for row in summary_sheet.iter_rows():
            for cell in row:
                cell.alignment = left_align
                cell.border = thin_border

        # Style Entries sheet
        # Headers
        for cell in entries_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Style all cells
        for row in entries_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border

        # Adjust column widths
        for sheet in [summary_sheet, entries_sheet]:
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook
        writer.close()

        # Send Excel file via Telegram
        await send_excel_file(file_path)
        excel_sent_date = today


# Main loop
async def main():
    while True:
        await process_entries()
        await asyncio.sleep(60)


if __name__ == "__main__":
    asyncio.run(main())
